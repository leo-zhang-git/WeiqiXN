# -*- coding: utf-8 -*-
"""
Excel类型检查器
提供Excel数据验证功能，支持校验表头格式和数据类型
"""

import re
import sys
from pathlib import Path
from typing import Any, Dict, List, Tuple, Callable, Optional

try:
    import openpyxl
except ImportError:
    raise ImportError("缺少必要的依赖库 openpyxl，请先运行 setup.bat")

# 导入检查器模块
from checker import BaseChecker, ColumnChecker, parse_extra_checkers
from checker.unique import UniqueChecker
from checker.enum import EnumChecker

# 高亮颜色
class Highlight:
    RED = '\033[91m'
    YELLOW = '\033[93m'
    GREEN = '\033[92m'
    CYAN = '\033[96m'
    BOLD = '\033[1m'
    END = '\033[0m'

    @classmethod
    def red(cls, text: str) -> str:
        return f"{cls.RED}{text}{cls.END}"

    @classmethod
    def yellow(cls, text: str) -> str:
        return f"{cls.YELLOW}{text}{cls.END}"

    @classmethod
    def green(cls, text: str) -> str:
        return f"{cls.GREEN}{text}{cls.END}"

    @classmethod
    def cyan(cls, text: str) -> str:
        return f"{cls.CYAN}{text}{cls.END}"

    @classmethod
    def bold(cls, text: str) -> str:
        return f"{cls.BOLD}{text}{cls.END}"

def col_num_to_excel(col_num: int) -> str:
    """将数字列号转换为Excel列号格式 (A, B, ..., Z, AA, AB, ...)"""
    result = ""
    while col_num > 0:
        col_num -= 1
        result = chr(65 + (col_num % 26)) + result
        col_num //= 26
    return result

# 基础类型验证函数
def _validate_float(v: Any) -> bool:
    """验证浮点数类型"""
    if isinstance(v, bool):
        return False
    if isinstance(v, (int, float)):
        return True
    try:
        float(v)
        return True
    except (ValueError, TypeError):
        return False

def _validate_int(v: Any) -> bool:
    """验证整数类型"""
    if isinstance(v, bool):
        return False
    if isinstance(v, int):
        return True
    try:
        float_val = float(v)
        return float_val == int(float_val)
    except (ValueError, TypeError):
        return False

# 基础类型转换函数
def _convert_float(v: Any) -> float:
    """转换为浮点数"""
    return float(v)

def _convert_int(v: Any) -> int:
    """转换为整数"""
    return int(float(v))

def _convert_string(v: Any) -> str:
    """转换为字符串"""
    return str(v)

def _convert_boolean(v: Any) -> bool:
    """转换为布尔值"""
    val_str = str(v).lower().strip()
    if val_str in ('true', '1', 'yes', 'on'):
        return True
    elif val_str in ('false', '0', 'no', 'off'):
        return False
    else:
        return bool(v)

# 基础类型名称集合
BASIC_TYPES = {'string', 'int', 'float', 'boolean'}

# 基础类型验证器映射
BASIC_VALIDATORS = {
    'string': lambda v: True,
    'float': _validate_float,
    'int': _validate_int,
    'boolean': lambda v: isinstance(v, bool),
}

# 基础类型转换器映射
BASIC_CONVERTERS = {
    'string': _convert_string,
    'float': _convert_float,
    'int': _convert_int,
    'boolean': _convert_boolean,
}

def _is_list_type(type_name: str) -> bool:
    """判断类型是否为列表类型 list(...)"""
    if not type_name.startswith('list(') or not type_name.endswith(')'):
        return False
    inner = type_name[5:-1]
    return inner in BASIC_TYPES

def _get_inner_type(type_name: str) -> str:
    """获取列表类型的内部元素类型"""
    return type_name[5:-1]

def _parse_list_value(value: str) -> List[str]:
    """解析列表字符串，返回元素列表"""
    value = value.strip()
    if not value.startswith('[') or not value.endswith(']'):
        raise ValueError("列表必须以[]包裹")
    content = value[1:-1].strip()
    if not content:
        return []
    return [elem.strip() for elem in content.split(',') if elem.strip()]

class ValidationError(Exception):
    """验证错误异常"""
    def __init__(self, row: int, col: int, message: str, sheet_name: str = ""):
        self.row = row
        self.col = col
        self.col_letter = col_num_to_excel(col)
        self.message = message
        self.sheet_name = sheet_name
        # 格式化带高亮的错误消息
        self.formatted = (
            f"[{Highlight.cyan(sheet_name)}] "
            f"第{Highlight.yellow(str(row))}行{Highlight.yellow(self.col_letter)}列: "
            f"{Highlight.red(message)}"
        )
        super().__init__(self.formatted)

class ExcelChecker:
    """Excel数据验证器"""

    TYPE_VALIDATORS = {
        'string': lambda v: True,
        'float': lambda v: _validate_float(v),
        'int': lambda v: _validate_int(v),
        'boolean': lambda v: isinstance(v, bool),
    }

    def __init__(self, excel_path: str | Path):
        self.excel_path = Path(excel_path)
        self.wb = None
        self.ws = None

    def load(self):
        if not self.excel_path.exists():
            raise FileNotFoundError(f"文件不存在: {self.excel_path}")
        self.wb = openpyxl.load_workbook(self.excel_path)
        self.ws = self.wb.active

    def close(self):
        if self.wb:
            self.wb.close()
            self.wb = None
            self.ws = None

    def get_valid_sheets(self) -> List[str]:
        if self.wb is None:
            raise RuntimeError("工作簿未加载")
        return [s for s in self.wb.sheetnames if not s.startswith('#')]

    def set_active_sheet(self, sheet_name: str):
        if self.wb is None:
            raise RuntimeError("工作簿未加载")
        if sheet_name in self.wb.sheetnames:
            self.ws = self.wb[sheet_name]
        else:
            raise ValueError(f"未找到工作表 '{sheet_name}'")

    def _get_cell_value(self, row: int, col: int) -> Any:
        return self.ws.cell(row=row, column=col).value

    def _is_valid_identifier(self, name: str) -> Tuple[bool, str]:
        if not name:
            return False, "key名不能为空"
        pattern = r'^[a-zA-Z_][a-zA-Z0-9_]*$'
        if not re.match(pattern, name):
            if name[0].isdigit():
                return False, f"key名 '{name}' 不能以数字开头"
            else:
                return False, f"key名 '{name}' 包含非法字符"
        return True, ""

    def _validate_type(self, value: Any, type_name: str) -> Tuple[bool, str]:
        """验证值是否符合指定类型"""
        if value is None or value == '':
            return False, f"值不能为空"

        # 处理列表类型
        if _is_list_type(type_name):
            return self._validate_list(value, type_name)

        # 处理基础类型
        if type_name not in BASIC_VALIDATORS:
            return False, f"不支持的数据类型: {type_name}"

        try:
            validator = BASIC_VALIDATORS[type_name]
            if not validator(value):
                return False, f"值 '{value}' 类型不匹配，期望 {type_name}"
            return True, ""
        except (ValueError, TypeError) as e:
            return False, f"值 '{value}' 无法转换为 {type_name}: {str(e)}"

    def _validate_list(self, value: Any, type_name: str) -> Tuple[bool, str]:
        """验证列表类型，调用基础类型验证器检查每个元素"""
        if not isinstance(value, str):
            return False, f"值 '{value}' 类型不匹配，期望 {type_name}"

        try:
            elements = _parse_list_value(value)
        except ValueError as e:
            return False, f"列表格式错误: {str(e)}"

        inner_type = _get_inner_type(type_name)
        for i, elem in enumerate(elements):
            # 使用基础类型验证器检查每个元素
            is_valid, error_msg = self._validate_type(elem, inner_type)
            if not is_valid:
                return False, f"列表第{i+1}个元素 {error_msg}"

        return True, ""

    def _convert_value(self, value: Any, type_name: str) -> Any:
        """将值按指定类型转换"""
        # 处理列表类型
        if _is_list_type(type_name):
            return self._convert_list(value, type_name)

        # 处理基础类型
        if type_name not in BASIC_CONVERTERS:
            return value

        return BASIC_CONVERTERS[type_name](value)

    def _convert_list(self, value: Any, type_name: str) -> List[Any]:
        """转换列表类型，调用基础类型转换器转换每个元素"""
        if not isinstance(value, str):
            raise ValueError(f"值必须是字符串类型")

        elements = _parse_list_value(value)
        inner_type = _get_inner_type(type_name)

        # 使用基础类型转换器转换每个元素
        return [self._convert_value(elem, inner_type) for elem in elements]

    def _find_value_row(self, col: int, target_value: str) -> Optional[int]:
        """查找指定值第一次出现的行号"""
        for row in range(5, self.ws.max_row + 1):
            value = self._get_cell_value(row, col)
            if value is not None and str(value).strip() == target_value:
                return row
        return None

    def _run_extra_checkers(self, col_values: Dict[str, List[Any]], headers: List[Dict], sheet_name: str):
        """执行列的特殊检查"""
        for header in headers:
            extra = header.get('extra', '')
            if not extra or extra == '#':
                continue

            checkers_info = parse_extra_checkers(extra)
            col = header['col']
            key = header['key']
            col_type = header.get('type', 'string')
            values = col_values.get(col, [])

            for func_name, args in checkers_info:
                checker_cls = ColumnChecker.get_checker(func_name)
                if checker_cls is None:
                    raise ValidationError(4, col, f"不支持的特殊检查: #{func_name}", sheet_name)

                passed, error_msg = checker_cls.check(values, col, key, sheet_name, col_type=col_type, args=args)

                if not passed:
                    if checker_cls.name == 'unique':
                        seen = {}
                        dup_value = None
                        for i, v in enumerate(values):
                            v_str = str(v).strip()
                            if v_str in seen:
                                dup_value = v_str
                                break
                            seen[v_str] = i
                        if dup_value:
                            first_row = self._find_value_row(col, dup_value)
                            if first_row:
                                raise ValidationError(first_row, col, error_msg, sheet_name)
                    else:
                        raise ValidationError(4, col, error_msg, sheet_name)

    def _is_valid_type(self, type_name: str) -> bool:
        """判断类型是否有效（基础类型或列表类型）"""
        return type_name in BASIC_TYPES or _is_list_type(type_name)

    def _get_supported_types(self) -> List[str]:
        """获取支持的所有类型"""
        return list(BASIC_TYPES) + ['list(...)']

    def parse_headers(self, sheet_name: str = "") -> List[Dict[str, Any]]:
        """解析表头配置"""
        if self.ws is None:
            raise RuntimeError("工作簿未加载")

        max_col = self.ws.max_column
        if max_col < 1:
            raise ValueError("Excel文件没有数据列")

        headers = []
        for col in range(1, max_col + 1):
            display_name = self._get_cell_value(1, col)
            key = self._get_cell_value(2, col)
            type_name = self._get_cell_value(3, col)
            extra = self._get_cell_value(4, col)

            if extra is None:
                raise ValidationError(4, col, "第4行不能为空", sheet_name)
            extra_str = str(extra).strip()
            if not extra_str.startswith('#'):
                raise ValidationError(4, col, f"第4行必须以#开头", sheet_name)

            if key is None or str(key).strip() == '':
                raise ValidationError(2, col, "key名不能为空", sheet_name)
            key_str = str(key).strip()
            is_valid, error_msg = self._is_valid_identifier(key_str)
            if not is_valid:
                raise ValidationError(2, col, error_msg, sheet_name)

            type_str = str(type_name).strip().lower() if type_name else ''
            if not self._is_valid_type(type_str):
                raise ValidationError(3, col,
                    f"不支持的数据类型 '{type_name}'，支持: {self._get_supported_types()}", sheet_name)

            # 验证特殊检查器
            checkers = parse_extra_checkers(extra_str)
            for func_name, args in checkers:
                checker = ColumnChecker.get_checker(func_name)
                if checker is None:
                    raise ValidationError(4, col, f"不支持的特殊检查: #{func_name}", sheet_name)
                if func_name == 'enum' and not args.strip():
                    raise ValidationError(4, col, "enum() 缺少参数，格式: #enum(a,b,c)", sheet_name)

            headers.append({
                'display_name': display_name,
                'key': key_str,
                'type': type_str,
                'extra': extra_str,
                'col': col
            })

        return headers

    def validate_data(self, headers: List[Dict[str, Any]], sheet_name: str = "") -> Dict[str, Dict[str, Any]]:
        """验证数据行的数据类型"""
        if self.ws is None:
            raise RuntimeError("工作簿未加载")

        max_row = self.ws.max_row
        data_dict = {}
        col_values = {}

        for header in headers:
            col_values[header['col']] = []

        for row in range(5, max_row + 1):
            first_col = headers[0]['col']
            first_value = self._get_cell_value(row, first_col)

            if first_value is None or str(first_value).strip() == '':
                raise ValidationError(row, first_col, "第一列（作为键）不能为空", sheet_name)

            first_value_str = str(first_value).strip()
            if first_value_str in data_dict:
                raise ValidationError(row, first_col, f"键 '{first_value_str}' 与之前行重复", sheet_name)

            row_data = {'_row': row}
            is_empty_row = True

            for header in headers:
                col = header['col']
                value = self._get_cell_value(row, col)

                # 收集所有值（包括空值），供检查器使用
                col_values[col].append(value)

                if value is None or str(value).strip() == '':
                    continue

                is_empty_row = False

                is_valid, error_msg = self._validate_type(value, header['type'])
                if not is_valid:
                    raise ValidationError(row, col, error_msg, sheet_name)

                try:
                    row_data[header['key']] = self._convert_value(value, header['type'])
                except (ValueError, TypeError) as e:
                    raise ValidationError(row, col, f"值转换失败: {str(e)}", sheet_name)

            if not is_empty_row:
                del row_data['_row']
                data_dict[first_value_str] = row_data

        self._run_extra_checkers(col_values, headers, sheet_name)
        return data_dict

    def check_sheet(self, sheet_name: str) -> Tuple[bool, str, int]:
        """检查单个工作表"""
        try:
            self.set_active_sheet(sheet_name)
            headers = self.parse_headers(sheet_name)
            data = self.validate_data(headers, sheet_name)
            return True, f"[{sheet_name}] 检查通过，共 {len(data)} 条数据", len(data)
        except ValidationError as e:
            return False, f"校验失败: {e}", 0
        except Exception as e:
            return False, f"[{sheet_name}] 未知错误: {str(e)}", 0

    def check_all(self) -> List[Tuple[str, bool, str, int]]:
        """检查所有有效工作表"""
        if self.wb is None:
            raise RuntimeError("工作簿未加载")

        results = []
        for sheet_name in self.get_valid_sheets():
            success, message, count = self.check_sheet(sheet_name)
            results.append((sheet_name, success, message, count))
        return results
